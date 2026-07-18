"""
===================================================================
 EXECUTIVE PDF + EXCEL REPORT GENERATOR (V8.3.0 — Master AI Trading Dashboard)
===================================================================
V8.3.0 REDESIGN (Task G6):
  PDF ko "Master AI Trading Dashboard" banaya gaya hai. Naya
  structure:
    1. Title header (large blue + subtitle + date/time + scanned count)
    2. 📊 Market Sentiment section (agar breadth dict pass hua)
       — compact advances/declines bar, signal distribution, sentiment
         label with emoji.
    3. 🤖 GLM AI Top Picks section (agar glm_picks list pass hui)
       — har pick ke liye highlighted box: rank, bold stock name,
         confidence meter (visual bar), action badge (🟢/🟡/🔴),
         Hinglish rationale, Hinglish risk note.
    4. 📈 Detailed Stock Analysis — per-stock card:
         • Header bar (emoji + #rank + bold stock name + signal badge)
         • Trade plan table (2-col, alternating rows, bold labels)
         • Chart image (thin border)
         • 💡 AI Analysis (Hinglish) — agar GLM pick available hai to
           🤖 GLM View subsection pehle dikhaya jaata hai.
         • 📰 Latest News (Hinglish) — bulleted.
    5. Disclaimer page (last) — Hinglish warning about technical
       analysis + do your own research.
  Footer har page par: page number + "AI Scanner V8.3.0 | Not
  financial advice" + generation timestamp.

  Excel: Scan Result sheet me Market Breadth info row top par add
  hoti hai (agar breadth available). Naya "GLM Picks" sheet ban
  gaya (rank, stock, name, confidence, action, rationale, risk).

V8.2.0 FIXES (preserved):
  1. PDF "AI Analysis" ab scanner.py ka REAL row['AI_Analysis']
     use karta hai (har stock ka alag, actual indicators par based).
  2. Entry/SL/target ab scanner.py ke row se aate hain (charts.py
     ke fib-swing se nahi) — Telegram/DB/PDF sab same numbers.
  3. News fetch shared news.py (cached) use karta hai.
  4. Hindi font cross-platform detection (Windows/Linux/Mac) +
     config.py PDF_HINDI_FONT_PATH override.
  5. Stock naam se ".NS" hata diya display mein.
  6. Excel report styled (color-coded signals + Top Buy List sheet).
  7. _safe_paragraph helper: ReportLab Paragraph XML-safe (V8.2.0
     bugs #8/#9 — dynamic text mein `<`, `>`, `&` aa sakta hai).
===================================================================
"""

import os
import re
from datetime import datetime, timezone, timedelta

import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    KeepTogether, PageBreak, HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from targets import calculate_targets
from utils import clean_symbol, escape_html
from news import format_news_text
from logger import logger

# ---------------------------------------------------------------
# V8.3.0: Hinglish translator (translator.py may have to_hinglish
# if G5 added it; else fall back to to_hindi; else English).
# ---------------------------------------------------------------
_HINGLISH_FN = None
try:
    from translator import to_hinglish as _to_hinglish_fn  # type: ignore
    _HINGLISH_FN = _to_hinglish_fn
except ImportError:
    try:
        from translator import to_hindi as _to_hindi_fn  # type: ignore
        _HINGLISH_FN = _to_hindi_fn
    except ImportError:
        _HINGLISH_FN = None


def _to_hinglish(text, max_len=400):
    """Hinglish text banata hai (preferred) -> Hindi (fallback) ->
    original English (last resort). Kabhi exception raise nahi karta."""
    if not text or not str(text).strip():
        return text
    if _HINGLISH_FN is None:
        return text
    try:
        # to_hinglish/to_hindi dono (text, max_len=) signature use karte hain
        return _HINGLISH_FN(text, max_len=max_len)
    except TypeError:
        # agar signature mismatch ho (koi future version arg change kare)
        try:
            return _HINGLISH_FN(text)
        except Exception:
            return text
    except Exception:
        return text


# V8.2.0 FIX (bugs #8, #9): ReportLab Paragraph XML-parser hai (HTML
# nahi). Dynamic text (stock names, AI analysis, news) mein agar `<`,
# `>`, ya `&` ho to PDF generate karte waqt ValueError raise hota
# tha. Safe-paragraph helper: pehle HTML-escape, phir Paragraph wrap.
def _safe_paragraph(text, style):
    """ReportLab Paragraph ke liye XML-safe text banata hai.
    Agar text None ho ya parse fail kare to safe fallback return."""
    if text is None:
        text = ""
    safe_text = escape_html(str(text))
    try:
        return Paragraph(safe_text, style)
    except Exception as e:
        logger.warning(f"PDF Paragraph parse fail (fallback plain text): {e}")
        return Paragraph(str(text).replace('<', '').replace('>', '').replace('&', ''), style)


try:
    from config import PDF_HINDI_FONT_PATH
except ImportError:
    PDF_HINDI_FONT_PATH = ""

try:
    from config import PDF_EMOJI_FONT_PATH
except ImportError:
    PDF_EMOJI_FONT_PATH = ""

# ---------------------------------------------------------------
# HINDI-CAPABLE FONT DETECTION (cross-platform)
# ---------------------------------------------------------------
_CANDIDATE_FONT_PATHS = [
    PDF_HINDI_FONT_PATH,  # user-specified override, config.py
    "C:\\Windows\\Fonts\\Nirmala.ttf",
    "C:\\Windows\\Fonts\\mangal.ttf",
    "C:\\Windows\\Fonts\\aparaj.ttf",
    "/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansDevanagari-Regular.ttf",
    "/System/Library/Fonts/Supplemental/DevanagariMT.ttc",
]

FONT_NAME = 'Helvetica'
for _path in _CANDIDATE_FONT_PATHS:
    if _path and os.path.exists(_path):
        try:
            pdfmetrics.registerFont(TTFont('HindiFont', _path))
            FONT_NAME = 'HindiFont'
            logger.info(f"PDF Hindi font mil gaya: {_path}")
            break
        except Exception:
            continue

if FONT_NAME == 'Helvetica':
    logger.warning(
        "Koi Devanagari-capable font nahi mila is machine par - PDF mein Hindi text "
        "boxes/blank dikh sakta hai. Fix: config.py mein PDF_HINDI_FONT_PATH set karo "
        "(apne system ka koi Hindi font path, e.g. Windows par 'C:\\Windows\\Fonts\\Nirmala.ttf')."
    )

# V8.3.0: HindiFont ko as a font family register karo (bold/italic
# variants same TTF ko point karte hain) — warna Paragraph mein
# <b> tag use karne par "font not found" warning aati thi. Ab
# <b> tag Hindi font ke saath bhi safe hai (regular weight render
# hoga, lekin crash nahi hoga).
if FONT_NAME == 'HindiFont':
    try:
        registerFontFamily(
            'HindiFont',
            normal='HindiFont', bold='HindiFont',
            italic='HindiFont', boldItalic='HindiFont',
        )
    except Exception:
        pass

# ---------------------------------------------------------------
# V8.3.0: FALLBACK FONT DETECTION — DejaVu Sans (better Unicode
# coverage than Helvetica). Hindi font nahi mila to DejaVu Sans
# try karte hain (Linux/Mac par commonly available, Windows par
# installable). Isse basic Latin + common symbols (• ₹ ━ ⚡ ⚠)
# saaf render hote hain. Supplementary plane emojis (🔥 💡 📰 etc.)
# ke liye alag emoji font chahiye (neeche dekho).
# ---------------------------------------------------------------
if FONT_NAME == 'Helvetica':
    _DEJAVU_CANDIDATES = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/DejaVuSans.ttf",
        "C:\\Windows\\Fonts\\dejavu.ttf",
        "C:\\Windows\\Fonts\\DejaVuSans.ttf",
    ]
    for _dpath in _DEJAVU_CANDIDATES:
        if _dpath and os.path.exists(_dpath):
            try:
                pdfmetrics.registerFont(TTFont('DejaVuSans', _dpath))
                # Bold variant bhi try karo (for <b> tag support)
                _dpath_bold = _dpath.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf")
                if os.path.exists(_dpath_bold):
                    try:
                        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', _dpath_bold))
                        registerFontFamily('DejaVuSans',
                                            normal='DejaVuSans', bold='DejaVuSans-Bold',
                                            italic='DejaVuSans', boldItalic='DejaVuSans-Bold')
                    except Exception:
                        registerFontFamily('DejaVuSans', normal='DejaVuSans',
                                            bold='DejaVuSans', italic='DejaVuSans',
                                            boldItalic='DejaVuSans')
                else:
                    registerFontFamily('DejaVuSans', normal='DejaVuSans',
                                        bold='DejaVuSans', italic='DejaVuSans',
                                        boldItalic='DejaVuSans')
                FONT_NAME = 'DejaVuSans'
                logger.info(f"PDF DejaVu Sans font mil gaya: {_dpath}")
                break
            except Exception:
                continue

# ---------------------------------------------------------------
# V8.3.0: EMOJI FONT DETECTION (best-effort) — supplementary plane
# emojis (🔥 💡 📰 🤖 📊 etc.) ke liye. Most emoji fonts (Noto Color
# Emoji, Apple Color Emoji, Segoe UI Emoji) color bitmap (CBDT/CBLC)
# use karte hain jo ReportLab TTFont load nahi kar paata. Symbola
# (monochrome) aur kuch others TTF hain. Agar koi emoji font mil
# gaya, use as EMOJI_FONT register karte hain; warna supplementary
# plane emojis PDF mein boxes (□) dikhenge — but structure intact.
# ---------------------------------------------------------------
EMOJI_FONT = None
_EMOJI_CANDIDATES = [
    PDF_EMOJI_FONT_PATH,  # user-specified override
    "C:\\Windows\\Fonts\\seguiemj.ttf",  # Segoe UI Emoji (Windows; likely color bitmap)
    "/System/Library/Fonts/Apple Color Emoji.ttc",  # Mac (color bitmap, may fail)
    "/usr/share/fonts/truetype/noto/NotoColorEmoji.ttf",  # Linux (color bitmap, may fail)
    "/usr/share/fonts/truetype/emoji/NotoColorEmoji.ttf",
    "/usr/share/fonts/truetype/symbola/Symbola.ttf",  # Symbola (monochrome, if installed)
    "/usr/share/fonts/Symbola.ttf",
]
for _epath in _EMOJI_CANDIDATES:
    if _epath and os.path.exists(_epath):
        try:
            pdfmetrics.registerFont(TTFont('EmojiFont', _epath))
            EMOJI_FONT = 'EmojiFont'
            logger.info(f"PDF emoji font mil gaya: {_epath}")
            break
        except Exception as _e:
            # Most color bitmap emoji fonts yahan fail karenge — silent skip
            logger.debug(f"Emoji font {_epath} load nahi ho paya: {_e}")
            continue

if EMOJI_FONT is None:
    logger.info(
        "Koi emoji-capable TTF font nahi mila. PDF mein supplementary plane emojis "
        "(🔥 💡 📰 🤖 📊 etc.) boxes (□) dikh sakte hain. Basic structure (colors, "
        "bold, tables) intact rahega. Fix: config.py mein PDF_EMOJI_FONT_PATH set karo."
    )

# Devanagari-capable fallback for analysis/news body text. Agar Hindi
# font available hai to body usse render hota hai (Devanagari glyphs
# saaf dikhenge). Nahi to DejaVu Sans (better Unicode) ya Helvetica.
BODY_FONT = FONT_NAME
# Bold variant: HindiFont -> same TTF (registered family); DejaVuSans
# -> DejaVuSans-Bold (registered family); Helvetica -> Helvetica-Bold.
if FONT_NAME == 'HindiFont':
    BODY_BOLD_FONT = 'HindiFont'
elif FONT_NAME == 'DejaVuSans':
    BODY_BOLD_FONT = 'DejaVuSans-Bold'
else:
    BODY_BOLD_FONT = 'Helvetica-Bold'


# ---------------------------------------------------------------
# V8.3.0 COLOR PALETTE (professional, dark blue + green/red accents)
# ---------------------------------------------------------------
COL_HEADER_BG = colors.HexColor('#0D47A1')     # dark blue (section bars)
COL_SUBHEAD_BG = colors.HexColor('#1565C0')    # medium blue (sub-bars)
COL_SECTION_BG = colors.HexColor('#F8F9FA')    # light grey (alt rows)
COL_CARD_BG = colors.HexColor('#FFFFFF')       # white card body
COL_BUY_GREEN = colors.HexColor('#2E7D32')     # green (BUY/targets)
COL_SL_RED = colors.HexColor('#C62828')        # red (SL/SELL)
COL_WATCH_AMBER = colors.HexColor('#F57F17')   # amber (WATCH/WAIT)
COL_TEXT_DARK = colors.HexColor('#263238')     # near-black
COL_TEXT_MUTED = colors.HexColor('#546E7A')    # grey
COL_BORDER_LIGHT = colors.HexColor('#E0E3EB')  # table grid
COL_DIVIDER = colors.HexColor('#CFD8DC')       # horizontal dividers
COL_GLM_HIGHLIGHT = colors.HexColor('#E3F2FD')  # light blue (GLM box)
COL_GLM_BORDER = colors.HexColor('#90CAF9')     # medium blue border
COL_BREADTH_HIGHLIGHT = colors.HexColor('#F1F8E9')  # light green (breadth box)

# IST timezone (footer timestamp ke liye)
_IST = timezone(timedelta(hours=5, minutes=30))


# ---------------------------------------------------------------
# V8.3.0 FOOTER (page number + disclaimer + timestamp)
# ---------------------------------------------------------------
class _FooterCallback:
    """Stateful footer drawer — har page par:
    • Left: "AI Scanner V8.3.0 | Page N"
    • Center: "Not financial advice — apna research zaroor karo"
    • Right: generation timestamp (IST)
    • Thin divider line footer ke upar.
    """

    def __init__(self, gen_ts):
        self.gen_ts = gen_ts or ""

    def __call__(self, canvas, doc):
        canvas.saveState()
        canvas.setFont(BODY_FONT, 7.5)
        canvas.setFillColor(COL_TEXT_MUTED)
        page_num = canvas.getPageNumber()
        footer_y = 12
        canvas.drawString(20, footer_y, f"AI Scanner V8.3.0  |  Page {page_num}")
        canvas.drawCentredString(
            letter[0] / 2.0, footer_y,
            "Not financial advice — apna research zaroor karo"
        )
        if self.gen_ts:
            canvas.drawRightString(letter[0] - 20, footer_y, self.gen_ts)
        # thin divider above footer
        canvas.setStrokeColor(COL_DIVIDER)
        canvas.setLineWidth(0.5)
        canvas.line(20, footer_y + 12, letter[0] - 20, footer_y + 12)
        canvas.restoreState()


# ---------------------------------------------------------------
# V8.3.0 HELPER: colored section header bar
# ---------------------------------------------------------------
def _section_header(title, emoji, bg_color=COL_HEADER_BG):
    """Full-width colored bar with emoji + bold white title."""
    style = ParagraphStyle(
        'SH', fontName=BODY_BOLD_FONT, fontSize=12,
        leading=16, textColor=colors.white, leftIndent=0,
    )
    p = Paragraph(f"{emoji} <b>{escape_html(title)}</b>", style)
    t = Table([[p]], colWidths=[letter[0] - 40])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), bg_color),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    return t


# ---------------------------------------------------------------
# V8.3.0 HELPER: signal badge (emoji + color + label)
# ---------------------------------------------------------------
def _signal_badge(signal):
    """Returns (emoji, color, label) for a Signal string."""
    s = (signal or "").upper()
    if s == "STRONG BUY":
        return "🔥", COL_BUY_GREEN, "STRONG BUY"
    if s == "BUY":
        return "⚡", COL_BUY_GREEN, "BUY"
    if s == "WATCH":
        return "👀", COL_WATCH_AMBER, "WATCH"
    if s in ("SELL / AVOID", "SELL", "AVOID"):
        return "⚠️", COL_SL_RED, "SELL / AVOID"
    return "•", COL_TEXT_MUTED, signal or "N/A"


def _glm_action_badge(action):
    """Returns (emoji, color, label) for GLM action string."""
    a = (action or "").upper()
    if a == "ENTER_NOW":
        return "🟢", COL_BUY_GREEN, "ENTER NOW"
    if a == "WAIT":
        return "🟡", COL_WATCH_AMBER, "WAIT"
    if a == "AVOID":
        return "🔴", COL_SL_RED, "AVOID"
    return "⚪", COL_TEXT_MUTED, action or "—"


# ---------------------------------------------------------------
# V8.3.0 HELPER: confidence meter (visual 10-cell bar)
# ---------------------------------------------------------------
def _confidence_bar(conf, max_val=10.0, cell_w=14.0, cell_h=10.0):
    """Visual confidence bar — 10 cells, filled cells colored green
    per score. Returns a small Table flowable (no text inside cells)."""
    try:
        score = float(conf) if conf is not None else 0.0
    except (TypeError, ValueError):
        score = 0.0
    score = max(0.0, min(max_val, score))
    filled = int(round(score / max_val * 10))
    row = [""] * 10
    t = Table([row], colWidths=[cell_w] * 10, rowHeights=[cell_h])
    style = [
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.white),
    ]
    for i in range(10):
        if i < filled:
            style.append(('BACKGROUND', (i, 0), (i, 0), COL_BUY_GREEN))
        else:
            style.append(('BACKGROUND', (i, 0), (i, 0), COL_BORDER_LIGHT))
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------
# V8.3.0 HELPER: advances vs declines mini-bar (10 cells)
# ---------------------------------------------------------------
def _ad_bar(advances, declines, cell_w=22.0, cell_h=12.0):
    """Advances (green) vs Declines (red) visual bar.
    Advances % decides kitne cells green hain."""
    total = (advances or 0) + (declines or 0)
    if total <= 0:
        pct = 50.0
    else:
        pct = (advances or 0) / total * 100.0
    filled_green = int(round(pct / 100.0 * 10))
    row = [""] * 10
    t = Table([row], colWidths=[cell_w] * 10, rowHeights=[cell_h])
    style = [
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.white),
    ]
    for i in range(10):
        if i < filled_green:
            style.append(('BACKGROUND', (i, 0), (i, 0), COL_BUY_GREEN))
        else:
            style.append(('BACKGROUND', (i, 0), (i, 0), COL_SL_RED))
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------
# V8.3.0 HELPER: wrap flowables in a single-cell colored box
# ---------------------------------------------------------------
def _colored_box(content_flowables, bg=COL_GLM_HIGHLIGHT, border=None,
                 col_width=None, pad=8):
    """Ek single-cell Table banata hai jisme `content_flowables`
    (list of Flowables) vertically stack hote hain. Background +
    optional border. Card-style visual ke liye use hota hai."""
    if not isinstance(content_flowables, list):
        content_flowables = [content_flowables]
    width = col_width or (letter[0] - 40)
    t = Table([[content_flowables]], colWidths=[width])
    style = [
        ('BACKGROUND', (0, 0), (-1, -1), bg),
        ('TOPPADDING', (0, 0), (-1, -1), pad),
        ('BOTTOMPADDING', (0, 0), (-1, -1), pad),
        ('LEFTPADDING', (0, 0), (-1, -1), pad + 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), pad + 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    if border is not None:
        style.append(('BOX', (0, 0), (-1, -1), 0.7, border))
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------
# V8.3.0 HELPER: news text (with bullets) -> Hinglish HTML for Paragraph
# ---------------------------------------------------------------
_NEWS_LINE_RE = re.compile(r"^([•\-\*]?\s*)(.+?)(\s*\([^)]+\))?$")

def _news_to_hinglish_html(news_text):
    """News text (with bullet lines like '• Title (Publisher)') ko
    Hinglish mein translate karke ReportLab Paragraph-safe HTML
    banata hai. Lines <br/> se join hoti hain. Bullets + publishers
    preserve karte hain, sirf title translate hota hai."""
    if not news_text or not str(news_text).strip():
        return "<i>Abhi koi recent news nahi mili.</i>"
    lines = [l.strip() for l in str(news_text).split("\n") if l.strip()]
    html_lines = []
    for line in lines:
        m = _NEWS_LINE_RE.match(line)
        if m:
            prefix, title, suffix = m.groups()
            translated = _to_hinglish(title, max_len=300)
            html_lines.append(
                f"{escape_html(prefix)}{escape_html(translated)}{escape_html(suffix or '')}"
            )
        else:
            translated = _to_hinglish(line, max_len=300)
            html_lines.append(escape_html(translated))
    return "<br/>".join(html_lines)


# ---------------------------------------------------------------
# V8.3.0 BUILD: title header (top of first page)
# ---------------------------------------------------------------
def _build_title_header(breadth=None):
    """Title header — large blue title + subtitle + date/time +
    scanned count. Top of first page (NOT a separate title page)."""
    now = datetime.now(_IST)
    date_str = now.strftime("%d %b %Y")
    time_str = now.strftime("%I:%M %p IST")

    title_style = ParagraphStyle(
        'DocTitle', fontName=BODY_BOLD_FONT, fontSize=22,
        leading=26, textColor=COL_HEADER_BG, alignment=TA_CENTER, spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', fontName=BODY_BOLD_FONT, fontSize=13,
        leading=16, textColor=COL_SUBHEAD_BG, alignment=TA_CENTER, spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        'DocMeta', fontName=BODY_FONT, fontSize=10,
        leading=14, textColor=COL_TEXT_MUTED, alignment=TA_CENTER,
    )

    flow = [
        Paragraph("🤖 AI STOCK SCANNER V8.3.0", title_style),
        Paragraph("Master AI Trading Dashboard", subtitle_style),
    ]

    # Scanned count line (if breadth available, use its total_scanned)
    if breadth and breadth.get("total_scanned"):
        scanned_str = f"Scanned <b>{breadth['total_scanned']}</b> stocks across NSE + BSE"
    else:
        scanned_str = "Scanned stocks across NSE + BSE"
    flow.append(
        Paragraph(
            f"📅 {escape_html(date_str)}  |  🕐 {escape_html(time_str)}  |  📊 {scanned_str}",
            meta_style,
        )
    )
    # Thin divider below title
    flow.append(Spacer(1, 4))
    flow.append(HRFlowable(width="100%", thickness=1.2, color=COL_HEADER_BG,
                            spaceBefore=2, spaceAfter=2))
    flow.append(Spacer(1, 6))
    return flow


# ---------------------------------------------------------------
# V8.3.0 BUILD: market sentiment section (compact, ~1/3 page)
# ---------------------------------------------------------------
def _build_breadth_section(breadth):
    """📊 Market Sentiment section — agar breadth dict pass hua.
    Compact layout: sentiment label (big, colored) on left,
    advances/declines bar + signal distribution on right."""
    if not breadth:
        return []

    sentiment = (breadth.get("sentiment") or "N/A").upper()
    emoji = breadth.get("sentiment_emoji") or "🟡"
    # Color by sentiment
    if sentiment == "BULLISH":
        sent_color = COL_BUY_GREEN
    elif sentiment == "BEARISH":
        sent_color = COL_SL_RED
    else:
        sent_color = COL_WATCH_AMBER

    sent_label_style = ParagraphStyle(
        'SentLabel', fontName=BODY_BOLD_FONT, fontSize=18,
        leading=22, textColor=sent_color, alignment=TA_CENTER,
    )
    sent_sub_style = ParagraphStyle(
        'SentSub', fontName=BODY_FONT, fontSize=9,
        leading=12, textColor=COL_TEXT_MUTED, alignment=TA_CENTER,
    )
    label_style = ParagraphStyle(
        'BL', fontName=BODY_BOLD_FONT, fontSize=9.5,
        leading=12, textColor=COL_TEXT_DARK,
    )
    value_style = ParagraphStyle(
        'BV', fontName=BODY_FONT, fontSize=9.5,
        leading=12, textColor=COL_TEXT_DARK,
    )

    # Left column: big sentiment label
    left_cell = [
        Paragraph(f"{emoji} <b>{escape_html(sentiment)}</b>", sent_label_style),
        Paragraph("Market Sentiment", sent_sub_style),
        Spacer(1, 6),
        Paragraph(
            f"<b>Breadth:</b> {breadth.get('breadth_pct', 0)}%",
            sent_sub_style,
        ),
    ]

    # Right column: advances/declines bar + stats grid
    adv = breadth.get("advances", 0)
    dec = breadth.get("declines", 0)
    ad_label = Paragraph(
        f"<b>Advances:</b> <font color='#2E7D32'>{adv}</font>  "
        f"|  <b>Declines:</b> <font color='#C62828'>{dec}</font>",
        value_style,
    )
    ad_bar = _ad_bar(adv, dec)

    # Stats: 2x3 grid
    stats_data = [
        [Paragraph("Total Scanned", label_style),
         Paragraph(str(breadth.get("total_scanned", 0)), value_style),
         Paragraph("New 20-day Highs", label_style),
         Paragraph(str(breadth.get("new_highs", 0)), value_style)],
        [Paragraph("🔥 Strong Buy", label_style),
         Paragraph(f"<font color='#2E7D32'><b>{breadth.get('strong_buy', 0)}</b></font>", value_style),
         Paragraph("⚡ Buy", label_style),
         Paragraph(f"<font color='#2E7D32'><b>{breadth.get('buy', 0)}</b></font>", value_style)],
        [Paragraph("👀 Watch", label_style),
         Paragraph(f"<font color='#F57F17'><b>{breadth.get('watch', 0)}</b></font>", value_style),
         Paragraph("⚠️ Sell/Avoid", label_style),
         Paragraph(f"<font color='#C62828'><b>{breadth.get('sell_avoid', 0)}</b></font>", value_style)],
    ]
    stats_t = Table(stats_data, colWidths=[80, 70, 100, 60])
    stats_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, COL_BORDER_LIGHT),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))

    right_cell = [
        ad_label,
        Spacer(1, 3),
        ad_bar,
        Spacer(1, 6),
        stats_t,
    ]

    # 2-col layout: left = sentiment, right = stats
    layout = Table(
        [[left_cell, right_cell]],
        colWidths=[(letter[0] - 40) * 0.32, (letter[0] - 40) * 0.68],
    )
    layout.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (0, 0), COL_BREADTH_HIGHLIGHT),
        ('BACKGROUND', (1, 0), (1, 0), COL_SECTION_BG),
        ('BOX', (0, 0), (-1, -1), 0.5, COL_BORDER_LIGHT),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, COL_BORDER_LIGHT),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))

    return [
        _section_header("MARKET SENTIMENT", "📊", bg_color=COL_HEADER_BG),
        Spacer(1, 4),
        layout,
        Spacer(1, 12),
    ]


# ---------------------------------------------------------------
# V8.3.0 BUILD: GLM AI Top Picks section
# ---------------------------------------------------------------
def _build_glm_picks_section(glm_picks):
    """🤖 GLM AI Top Picks section — agar glm_picks list pass hui.
    Har pick ke liye highlighted box: rank badge, bold stock name,
    confidence meter, action badge, Hinglish rationale + risk note."""
    if not glm_picks:
        return []

    name_style = ParagraphStyle(
        'GLMName', fontName=BODY_BOLD_FONT, fontSize=13,
        leading=16, textColor=COL_HEADER_BG,
    )
    meta_style = ParagraphStyle(
        'GLMMeta', fontName=BODY_FONT, fontSize=9.5,
        leading=12, textColor=COL_TEXT_MUTED,
    )
    action_style = ParagraphStyle(
        'GLMAction', fontName=BODY_BOLD_FONT, fontSize=10,
        leading=13, textColor=COL_TEXT_DARK,
    )
    body_style = ParagraphStyle(
        'GLMBody', fontName=BODY_FONT, fontSize=9.5,
        leading=13, textColor=COL_TEXT_DARK, leftIndent=2,
    )
    risk_style = ParagraphStyle(
        'GLMRisk', fontName=BODY_FONT, fontSize=9.5,
        leading=13, textColor=COL_SL_RED, leftIndent=2,
    )
    conf_label_style = ParagraphStyle(
        'GLMConf', fontName=BODY_FONT, fontSize=8.5,
        leading=11, textColor=COL_TEXT_MUTED,
    )

    pick_cards = []
    for p in glm_picks:
        sym_raw = p.get("stock", "")
        sym = clean_symbol(sym_raw)
        rank = p.get("rank", "?")
        name = p.get("name", "") or ""
        conf = p.get("glm_confidence")
        action = p.get("glm_action", "") or ""
        rationale = p.get("glm_rationale", "") or ""
        risk = p.get("glm_risk_note", "") or ""
        tech_score = p.get("technical_score", 0)

        action_emoji, action_color, action_label = _glm_action_badge(action)

        # Top row: "#rank SYM (Name) | Action: 🟢 ENTER NOW"
        header_html = (
            f"<b>#{escape_html(str(rank))}  {escape_html(sym)}</b>"
            + (f"  <font size='9' color='#546E7A'>({escape_html(name)})</font>" if name else "")
        )
        header_p = Paragraph(header_html, name_style)

        action_p = Paragraph(
            f"<font color='{action_color.hexval()}'><b>{action_emoji} {escape_html(action_label)}</b></font>",
            action_style,
        )

        # Top header table: name (left) | action (right)
        header_t = Table(
            [[header_p, action_p]],
            colWidths=[(letter[0] - 40 - 20) * 0.65, (letter[0] - 40 - 20) * 0.35],
        )
        header_t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))

        # Confidence meter row
        conf_str = f"{conf}/10" if conf is not None else "N/A"
        conf_meter = _confidence_bar(conf)
        conf_text_p = Paragraph(
            f"🎯 Confidence: <b>{escape_html(conf_str)}</b>  "
            f"|  📊 Tech Score: <b>{tech_score}/100</b>",
            meta_style,
        )
        # Confidence meter + text in a row
        conf_t = Table(
            [[conf_text_p, conf_meter]],
            colWidths=[(letter[0] - 40 - 20) * 0.70, (letter[0] - 40 - 20) * 0.30],
        )
        conf_t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))

        # Rationale
        rationale_p = _safe_paragraph(
            f"💡 <b>Rationale:</b> {rationale}" if rationale else "💡 <b>Rationale:</b> N/A",
            body_style,
        )

        # Risk
        if risk:
            risk_p = _safe_paragraph(f"⚠️ <b>Risk:</b> {risk}", risk_style)
        else:
            risk_p = Paragraph("⚠️ <b>Risk:</b> N/A", risk_style)

        # Combine all into a highlighted box
        box_content = [
            header_t,
            Spacer(1, 4),
            conf_t,
            Spacer(1, 4),
            rationale_p,
            Spacer(1, 2),
            risk_p,
        ]
        # Col width slightly narrower than full to leave padding inside _colored_box
        card = _colored_box(
            box_content, bg=COL_GLM_HIGHLIGHT, border=COL_GLM_BORDER, pad=8,
        )
        pick_cards.append(card)
        pick_cards.append(Spacer(1, 6))

    # Footer note
    note_style = ParagraphStyle(
        'GLMNote', fontName=BODY_FONT, fontSize=8.5,
        leading=11, textColor=COL_TEXT_MUTED, alignment=TA_CENTER,
    )
    note = Paragraph(
        "<i>🧠 GLM AI ne technical indicators dekh kar ye picks chune hain. "
        "Apna research zaroor karo.</i>",
        note_style,
    )

    return [
        _section_header("GLM AI TOP PICKS", "🤖", bg_color=COL_HEADER_BG),
        Spacer(1, 4),
    ] + pick_cards + [note, Spacer(1, 12)]


# ---------------------------------------------------------------
# V8.3.0 BUILD: per-stock detailed analysis card
# ---------------------------------------------------------------
def _build_stock_card(idx, data, glm_pick_for_stock=None):
    """Per-stock detailed analysis card:
      1. Header bar (emoji + #rank + bold stock name + signal badge)
      2. Trade plan table (2-col, alternating rows, bold labels)
      3. Chart image (thin border)
      4. 💡 AI Analysis (Hinglish) + 🤖 GLM View subsection (if available)
      5. 📰 Latest News (Hinglish) — bulleted.
    Returns a list of flowables (wrapped in KeepTogether at caller).
    """
    sym = data.get('symbol', '?')
    sym_safe = escape_html(sym)
    score = data.get('score', 0)
    signal = data.get('signal', '')
    signal_emoji, signal_color, signal_label = _signal_badge(signal)

    # ---- Header bar (colored, white text) ----
    header_style = ParagraphStyle(
        'CardHdr', fontName=BODY_BOLD_FONT, fontSize=14,
        leading=18, textColor=colors.white,
    )
    header_sub_style = ParagraphStyle(
        'CardHdrSub', fontName=BODY_FONT, fontSize=10,
        leading=13, textColor=colors.white,
    )
    # Left: 🔥 #1 RELIANCE   Right: STRONG BUY (badge style)
    left_html = f"{signal_emoji} <b>#{idx}  {sym_safe}</b>"
    right_html = f"<b>{escape_html(signal_label)}</b>"
    header_t = Table(
        [[Paragraph(left_html, header_style),
          Paragraph(right_html, header_style)]],
        colWidths=[(letter[0] - 40 - 16) * 0.65, (letter[0] - 40 - 16) * 0.35],
    )
    header_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COL_HEADER_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    # Score + indicators sub-bar
    rsi = data.get('rsi', '—')
    adx = data.get('adx', '—')
    rr = data.get('risk_reward', '—')
    close = data.get('close', 0.0)
    meta_html = (
        f"<b>Score:</b> {score}/100  "
        f"|  <b>CMP:</b> ₹{close:.2f}  "
        f"|  <b>R:R:</b> 1:{escape_html(str(rr))}  "
        f"|  <b>RSI:</b> {escape_html(str(rsi))}  "
        f"|  <b>ADX:</b> {escape_html(str(adx))}"
    )
    meta_t = Table(
        [[Paragraph(meta_html, header_sub_style)]],
        colWidths=[letter[0] - 40 - 16],
    )
    meta_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COL_SUBHEAD_BG),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    # ---- Trade plan table (2-col, alternating rows, bold labels) ----
    label_style = ParagraphStyle(
        'TPLbl', fontName=BODY_BOLD_FONT, fontSize=10,
        leading=13, textColor=COL_TEXT_DARK,
    )
    value_style = ParagraphStyle(
        'TPVal', fontName=BODY_FONT, fontSize=10,
        leading=13, textColor=COL_TEXT_DARK,
    )
    sl_value_style = ParagraphStyle(
        'TPSL', fontName=BODY_FONT, fontSize=10,
        leading=13, textColor=COL_SL_RED,
    )
    target_value_style = ParagraphStyle(
        'TPtgt', fontName=BODY_FONT, fontSize=10,
        leading=13, textColor=COL_BUY_GREEN,
    )

    def _fmt_price(v):
        try:
            return f"₹{float(v):.2f}"
        except (TypeError, ValueError):
            return str(v) if v is not None else "—"

    entry_zone_str = (
        f"{_fmt_price(data.get('entry_low'))} - {_fmt_price(data.get('entry_high'))}"
        if data.get('entry_low') and data.get('entry_high')
        else _fmt_price(data.get('entry'))
    )

    rows = [
        ("Entry Zone", entry_zone_str, value_style),
        ("Stop Loss", _fmt_price(data.get('sl')), sl_value_style),
        ("Target 1", _fmt_price(data.get('t1')), target_value_style),
        ("Target 2", _fmt_price(data.get('t2')), target_value_style),
        ("Final Target", _fmt_price(data.get('t3')), target_value_style),
        ("Support", _fmt_price(data.get('support')), value_style),
        ("Resistance", _fmt_price(data.get('resistance')), value_style),
        ("Pattern", escape_html(data.get('pattern', '—')), value_style),
        ("Weekly Trend", escape_html(data.get('weekly_trend', 'N/A')), value_style),
        ("1H Confirmation", escape_html(data.get('mtf_status', 'N/A')), value_style),
    ]

    table_data = [
        [Paragraph(label, label_style), Paragraph(str(value), vstyle)]
        for label, value, vstyle in rows
    ]
    trade_t = Table(table_data, colWidths=[150, letter[0] - 40 - 16 - 150])
    trade_style = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.3, COL_BORDER_LIGHT),
    ]
    # Alternating row backgrounds
    for r_idx in range(len(rows)):
        if r_idx % 2 == 0:
            trade_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), COL_SECTION_BG))
        else:
            trade_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.white))
    trade_t.setStyle(TableStyle(trade_style))

    # ---- Chart image (thin border) ----
    chart_path = data.get('chart_path')
    chart_flow = []
    if chart_path and os.path.exists(chart_path):
        try:
            # 500x300 (aspect 5:3) — leaves breathing room on page
            img = Image(chart_path, width=500, height=300)
            img_t = Table([[img]], colWidths=[letter[0] - 40 - 16])
            img_t.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, COL_BORDER_LIGHT),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ]))
            chart_flow = [img_t]
        except Exception as e:
            logger.warning(f"PDF Image load fail for {chart_path}: {e}")
            chart_flow = []
    else:
        chart_flow = [
            Paragraph(
                "<i>Chart image available nahi.</i>",
                ParagraphStyle('NoChart', fontName=BODY_FONT, fontSize=9,
                                leading=12, textColor=COL_TEXT_MUTED, alignment=TA_CENTER),
            )
        ]

    # ---- AI Analysis section ----
    analysis_header_style = ParagraphStyle(
        'AH', fontName=BODY_BOLD_FONT, fontSize=11,
        leading=15, textColor=COL_HEADER_BG, spaceBefore=2, spaceAfter=2,
    )
    analysis_body_style = ParagraphStyle(
        'AB', fontName=BODY_FONT, fontSize=10,
        leading=14, textColor=COL_TEXT_DARK, leftIndent=4,
    )
    glm_body_style = ParagraphStyle(
        'GB', fontName=BODY_FONT, fontSize=9.5,
        leading=13, textColor=COL_TEXT_DARK, leftIndent=4,
    )
    glm_risk_style = ParagraphStyle(
        'GR', fontName=BODY_FONT, fontSize=9.5,
        leading=13, textColor=COL_SL_RED, leftIndent=4,
    )

    analysis_flow = [Paragraph("💡 <b>AI Analysis (Hinglish):</b>", analysis_header_style)]
    ai_text = data.get('ai_analysis') or "Analysis available nahi hai."
    analysis_flow.append(_safe_paragraph(ai_text, analysis_body_style))

    # ---- GLM View subsection (if this stock has a GLM pick) ----
    if glm_pick_for_stock:
        glm_conf = glm_pick_for_stock.get("glm_confidence")
        glm_action = glm_pick_for_stock.get("glm_action", "") or ""
        glm_rat = glm_pick_for_stock.get("glm_rationale", "") or ""
        glm_risk = glm_pick_for_stock.get("glm_risk_note", "") or ""
        action_emoji, action_color, action_label = _glm_action_badge(glm_action)

        glm_hdr_html = (
            f"🤖 <b>GLM View:</b> "
            f"<font color='{action_color.hexval()}'><b>{action_emoji} {escape_html(action_label)}</b></font>"
            + (f"  |  🎯 Confidence: <b>{glm_conf}/10</b>" if glm_conf is not None else "")
        )
        glm_box_content = [
            Paragraph(glm_hdr_html, analysis_header_style),
        ]
        if glm_rat:
            glm_box_content.append(_safe_paragraph(
                f"<b>Rationale:</b> {glm_rat}", glm_body_style,
            ))
        if glm_risk:
            glm_box_content.append(_safe_paragraph(
                f"<b>Risk:</b> {glm_risk}", glm_risk_style,
            ))
        glm_box = _colored_box(
            glm_box_content, bg=COL_GLM_HIGHLIGHT, border=COL_GLM_BORDER, pad=6,
        )
        analysis_flow.append(Spacer(1, 4))
        analysis_flow.append(glm_box)

    # ---- News section ----
    news_header_style = ParagraphStyle(
        'NH', fontName=BODY_BOLD_FONT, fontSize=11,
        leading=15, textColor=COL_HEADER_BG, spaceBefore=6, spaceAfter=2,
    )
    news_body_style = ParagraphStyle(
        'NB', fontName=BODY_FONT, fontSize=9.5,
        leading=13, textColor=COL_TEXT_DARK, leftIndent=4,
    )
    news_flow = [
        Paragraph("📰 <b>Latest News (Hinglish):</b>", news_header_style),
    ]
    news_html = _news_to_hinglish_html(data.get('news'))
    try:
        news_flow.append(Paragraph(news_html, news_body_style))
    except Exception:
        # last-resort fallback
        news_flow.append(_safe_paragraph(
            (data.get('news') or "").replace("\n", " "), news_body_style,
        ))

    # ---- Combine into a card ----
    # NOTE: Outer 1-cell Table wrap use nahi karte — Tables apne
    # cells ko across pages split nahi karte, aur ek full card
    # (header + meta + trade table + chart + analysis + GLM view +
    # news) easily 700+ points tall ho jaata hai > ek page (744pt).
    # ReportLab LayoutError throw karta tha. Instead, chhote logical
    # chunks ko KeepTogether se bandhte hain; baki flowables naturally
    # across pages flow karte hain. Header bar (dark blue) hi visual
    # "card start" marker hai.
    flow = []

    # Chunk 1: header bar + meta bar (compact, must stay together)
    flow.append(KeepTogether([header_t, meta_t, Spacer(1, 6)]))

    # Chunk 2: trade plan table (atomic Table flowable — naturally
    # moves to next page if doesn't fit)
    flow.append(trade_t)
    flow.append(Spacer(1, 6))

    # Chunk 3: chart image (atomic Image flowable — naturally moves
    # to next page if doesn't fit)
    flow.extend(chart_flow)
    flow.append(Spacer(1, 6))

    # Chunk 4: AI analysis (header + first paragraph together — avoid
    # orphan "💡 AI Analysis" header at bottom of page)
    flow.append(KeepTogether(analysis_flow))
    flow.append(Spacer(1, 2))

    # Chunk 5: news (header + content together)
    flow.append(KeepTogether(news_flow))

    # Visual separator between cards (thin divider)
    flow.append(Spacer(1, 4))
    flow.append(HRFlowable(width="100%", thickness=0.5, color=COL_DIVIDER,
                            spaceBefore=2, spaceAfter=10))
    return flow


# ---------------------------------------------------------------
# V8.3.0 BUILD: disclaimer page (last)
# ---------------------------------------------------------------
def _build_disclaimer():
    """Hinglish disclaimer page — technical analysis, not financial
    advice, do your own research."""
    title_style = ParagraphStyle(
        'DiscTitle', fontName=BODY_BOLD_FONT, fontSize=16,
        leading=20, textColor=COL_HEADER_BG, alignment=TA_CENTER, spaceAfter=10,
    )
    body_style = ParagraphStyle(
        'DiscBody', fontName=BODY_FONT, fontSize=10.5,
        leading=15, textColor=COL_TEXT_DARK, alignment=TA_LEFT,
        leftIndent=20, rightIndent=20, spaceAfter=8,
    )
    bullet_style = ParagraphStyle(
        'DiscBul', fontName=BODY_FONT, fontSize=10,
        leading=14, textColor=COL_TEXT_DARK, alignment=TA_LEFT,
        leftIndent=40, rightIndent=20, spaceAfter=4,
    )

    return [
        PageBreak(),
        _section_header("DISCLAIMER", "⚠️", bg_color=COL_SL_RED),
        Spacer(1, 14),
        Paragraph("⚠️ <b>Important Disclaimer — Dhyan se padho</b>", title_style),
        Paragraph(
            "Ye report sirf <b>technical analysis</b> par based hai. "
            "Stock market mein investment  risk ke saath aata hai. "
            "Pehle se koi bhi return guarantee nahi hoti.",
            body_style,
        ),
        Paragraph("📍 <b>Key points:</b>", body_style),
        Paragraph(
            "• Ye report <b>financial advice nahi hai</b> — sirf "
            "educational aur informational purpose ke liye hai.", bullet_style),
        Paragraph(
            "• AI (GLM) aur rule-based scanner dono ne technical "
            "indicators (RSI, MACD, ADX, EMA, volume) dekh kar ye "
            "picks chune hain. Fundamentals (earnings, debt, management) "
            "ka analysis yahan nahi hua.", bullet_style),
        Paragraph(
            "• Past performance future returns ka guarantee nahi. "
            "Chart patterns kabhi bhi fail ho sakte hain.", bullet_style),
        Paragraph(
            "• Stop Loss hamesha apni risk capacity ke hisaab se "
            "decide karo. 1-2% risk per trade se zyada mat lo.", bullet_style),
        Paragraph(
            "• Targets aur Entry zones sirf suggestions hain — "
            "apne broker ke slippage + charges alag se calculate karo.", bullet_style),
        Paragraph(
            "• Hinglish translation automatic hai (Google Translate) — "
            "kabhi-kabhi meaning clear nahi ho sakta. Original English "
            "news padhne ke liye source check karo.", bullet_style),
        Spacer(1, 14),
        Paragraph(
            "<b>Apna research zaroor karo. Apne financial advisor se "
            "consult karo before making any investment decision.</b>",
            body_style,
        ),
        Spacer(1, 18),
        HRFlowable(width="60%", thickness=0.8, color=COL_DIVIDER,
                    spaceBefore=4, spaceAfter=4),
        Paragraph(
            "<i>Generated by AI Stock Scanner V8.3.0 — Master AI Trading Dashboard</i>",
            ParagraphStyle('DiscFoot', fontName=BODY_FONT, fontSize=9,
                            leading=12, textColor=COL_TEXT_MUTED, alignment=TA_CENTER),
        ),
    ]


# ---------------------------------------------------------------
# V9.4: FUNDAMENTAL CHECKLIST REPORT SECTION
# ---------------------------------------------------------------
def _build_fundamental_checklist_section(top_stocks_data):
    """
    V9.4: Fundamental checklist report for all BUY signal stocks.
    Shows pass/fail per condition + overall rating.
    """
    story = []
    story.append(_section_header("FUNDAMENTAL CHECKLIST REPORT", "📊"))
    story.append(Spacer(1, 8))

    # Intro text
    intro_style = ParagraphStyle(
        'ChecklistIntro', fontName=BODY_FONT, fontSize=10,
        leading=14, textColor=COL_TEXT_MUTED, alignment=TA_JUSTIFY,
    )
    story.append(Paragraph(
        "Is section mein har BUY signal stock ka fundamental verification hai. "
        "10 conditions check ki gayi hain — P/E, Debt, Book Value, Quarterly/Yearly "
        "Growth, ROE. Rating A+ (excellent) se D (weak) tak.",
        intro_style,
    ))
    story.append(Spacer(1, 12))

    try:
        from fundamental_analyzer import FundamentalAnalyzer, format_fundamental_checklist
        analyzer = FundamentalAnalyzer()

        for data in top_stocks_data[:10]:  # max 10 stocks
            symbol = data.get('symbol', '')
            if not symbol:
                continue

            # Add .NS suffix if not present
            if not symbol.endswith('.NS') and not symbol.endswith('.BO'):
                symbol = f"{symbol}.NS"

            try:
                result = analyzer.analyze(symbol)
                result_dict = result.to_dict()

                # Build checklist table for this stock
                story.extend(_build_single_fundamental_checklist(symbol, result_dict))
                story.append(Spacer(1, 10))

            except Exception as e:
                logger.warning(f"Fundamental checklist fail {symbol}: {e}")
                story.append(Paragraph(
                    f"<b>{symbol}</b>: Fundamental data fetch fail ({e})",
                    intro_style,
                ))
                story.append(Spacer(1, 6))

    except ImportError:
        story.append(Paragraph(
            "<i>Fundamental Analyzer module not available.</i>",
            intro_style,
        ))

    return story


def _build_single_fundamental_checklist(symbol, result_dict):
    """Build a single stock's fundamental checklist table."""
    from utils import clean_symbol, escape_html as _esc

    flowables = []
    display = _esc(clean_symbol(symbol))
    rating = result_dict.get("overall_rating", "N/A")
    score = result_dict.get("fundamental_score", 0)
    pass_count = result_dict.get("pass_count", 0)
    fail_count = result_dict.get("fail_count", 0)
    company = _esc(result_dict.get("company_name", ""))

    # Rating color
    rating_colors = {
        "A+": colors.HexColor("#2E7D32"),
        "A": colors.HexColor("#388E3C"),
        "B": colors.HexColor("#F57F17"),
        "C": colors.HexColor("#E65100"),
        "D": colors.HexColor("#C62828"),
        "N/A": colors.HexColor("#757575"),
    }
    rating_color = rating_colors.get(rating, colors.HexColor("#757575"))

    # Stock header with rating
    header_style = ParagraphStyle(
        'FundHeader', fontName=BODY_FONT_BOLD, fontSize=13,
        leading=16, textColor=colors.white, alignment=TA_LEFT,
    )
    header_data = [[
        Paragraph(f"<b>{display}</b> — {company}", header_style),
        Paragraph(f'<font color="white"><b>Rating: {rating}</b></font>',
                  ParagraphStyle('Rate', fontName=BODY_FONT_BOLD, fontSize=14,
                                 leading=18, textColor=colors.white, alignment=TA_CENTER)),
    ]]
    header_tbl = Table(header_data, colWidths=[400, 130])
    header_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), COL_HEADER_BG),
        ('BACKGROUND', (1, 0), (1, 0), rating_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    flowables.append(header_tbl)
    flowables.append(Spacer(1, 4))

    # Summary line
    summary_style = ParagraphStyle(
        'FundSummary', fontName=BODY_FONT, fontSize=9,
        leading=12, textColor=COL_TEXT_MUTED,
    )
    flowables.append(Paragraph(
        f"Score: {score}/100 | Pass: {pass_count} | Fail: {fail_count}",
        summary_style,
    ))
    flowables.append(Spacer(1, 6))

    # Checklist table
    check_style = ParagraphStyle(
        'CheckCell', fontName=BODY_FONT, fontSize=9,
        leading=12, textColor=COL_TEXT,
    )
    check_bold = ParagraphStyle(
        'CheckBold', fontName=BODY_FONT_BOLD, fontSize=9,
        leading=12, textColor=COL_TEXT,
    )

    # Table header
    table_data = [[
        Paragraph("<b>Check</b>", ParagraphStyle('TH', fontName=BODY_FONT_BOLD,
                   fontSize=9, leading=12, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Condition</b>", ParagraphStyle('TH', fontName=BODY_FONT_BOLD,
                   fontSize=9, leading=12, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Value</b>", ParagraphStyle('TH', fontName=BODY_FONT_BOLD,
                   fontSize=9, leading=12, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Result</b>", ParagraphStyle('TH', fontName=BODY_FONT_BOLD,
                   fontSize=9, leading=12, textColor=colors.white, alignment=TA_CENTER)),
    ]]

    for check in result_dict.get("checks", []):
        name = _esc(check.get("name", ""))
        condition = _esc(check.get("condition", ""))
        value = _esc(str(check.get("value", "")))
        passed = check.get("passed", False)

        if value == "N/A":
            result_text = "N/A"
            result_color = colors.HexColor("#9E9E9E")
        elif passed:
            result_text = "PASS"
            result_color = colors.HexColor("#2E7D32")
        else:
            result_text = "FAIL"
            result_color = colors.HexColor("#C62828")

        result_para = Paragraph(
            f'<font color="{result_color.hexval()}"><b>{result_text}</b></font>',
            ParagraphStyle('Res', fontName=BODY_FONT_BOLD, fontSize=10,
                           leading=13, alignment=TA_CENTER),
        )

        table_data.append([
            Paragraph(name, check_bold),
            Paragraph(condition, check_style),
            Paragraph(value, check_style),
            result_para,
        ])

    if len(table_data) > 1:
        checklist_tbl = Table(table_data, colWidths=[140, 130, 120, 60])
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), COL_HEADER_BG),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.3, COL_BORDER),
        ]
        # Alternating row colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F5F5F5")))
        checklist_tbl.setStyle(TableStyle(style_cmds))
        flowables.append(checklist_tbl)

    flowables.append(Spacer(1, 8))
    return flowables


# ---------------------------------------------------------------
# MAIN PDF GENERATOR (V8.3.0 — Master AI Trading Dashboard)
# ---------------------------------------------------------------
def generate_pdf_report(top_stocks_data, output_path, glm_picks=None, breadth=None):
    """
    V8.3.0 Master AI Trading Dashboard PDF generator.

    Params:
      top_stocks_data: list of per-stock dicts (from save_report)
      output_path: PDF file path
      glm_picks: optional list of GLM pick dicts (from glm_screener)
      breadth: optional breadth dict (from market_breadth)

    Page flow:
      1. Title header (top of first page)
      2. 📊 Market Sentiment (if breadth)
      3. 🤖 GLM AI Top Picks (if glm_picks)
      4. 📈 Detailed Stock Analysis (per-stock cards)
      5. ⚠️ Disclaimer page (last)
    """
    # Generation timestamp (footer ke liye)
    gen_ts = datetime.now(_IST).strftime("%d %b %Y, %I:%M %p IST")

    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=24,
        title="AI Stock Scanner V8.3.0 — Master AI Trading Dashboard",
        author="AI Stock Scanner V8.3.0",
    )
    story = []

    # ---- 1. Title header ----
    story.extend(_build_title_header(breadth=breadth))

    # ---- 2. Market Sentiment section (if breadth) ----
    story.extend(_build_breadth_section(breadth))

    # ---- 3. GLM AI Top Picks section (if glm_picks) ----
    story.extend(_build_glm_picks_section(glm_picks))

    # ---- 4. Detailed Stock Analysis ----
    if not top_stocks_data:
        empty_style = ParagraphStyle(
            'Empty', fontName=BODY_FONT, fontSize=11,
            leading=15, textColor=COL_TEXT_MUTED, alignment=TA_CENTER,
        )
        story.append(_section_header("DETAILED STOCK ANALYSIS", "📈"))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            "Aaj koi BUY / STRONG BUY signal nahi mila. Kal phir try karo.",
            empty_style,
        ))
    else:
        # GLM picks lookup by symbol (clean_symbol form) for per-stock cards
        glm_lookup = {}
        if glm_picks:
            for p in glm_picks:
                key = clean_symbol(p.get("stock", ""))
                if key:
                    glm_lookup[key] = p

        story.append(_section_header("DETAILED STOCK ANALYSIS", "📈"))
        story.append(Spacer(1, 8))

        for idx, data in enumerate(top_stocks_data, 1):
            sym = data.get('symbol', '')
            glm_pick = glm_lookup.get(sym)
            story.extend(_build_stock_card(idx, data, glm_pick_for_stock=glm_pick))

    # ---- 4.5 Fundamental Checklist Report (V9.4) ----
    if top_stocks_data:
        story.append(PageBreak())
        story.extend(_build_fundamental_checklist_section(top_stocks_data))

    # ---- 5. Disclaimer page ----
    story.extend(_build_disclaimer())

    # Build with footer on every page
    footer = _FooterCallback(gen_ts)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


# ---------------------------------------------------------------
# EXCEL REPORT (styled, V8.3.0 — GLM Picks sheet + Breadth info row)
# ---------------------------------------------------------------
EXCEL_COLUMNS = [
    "Stock", "Signal", "Score", "Close", "RSI", "ADX", "Supertrend", "Weekly_Trend",
    "Entry_Low", "Entry_High", "Stoploss", "Target_1", "Target_2", "Final_Target",
    "Support", "Resistance", "Risk_Reward", "Patterns",
]

SIGNAL_COLORS = {
    "STRONG BUY": "1B5E20",
    "BUY": "66BB6A",
    "WATCH": "FFEB3B",
    "SELL / AVOID": "EF5350",
}

GLM_ACTION_COLORS = {
    "ENTER_NOW": "C8E6C9",  # light green
    "WAIT": "FFF9C4",        # light yellow
    "AVOID": "FFCDD2",       # light red
}


def _prep_excel_dataframe(ranked_result):
    rows = []
    for r in ranked_result:
        entry, sl = r.get("Entry"), r.get("Stoploss")
        t1, t2, t3 = calculate_targets(entry, sl) if (entry and sl) else (None, None, None)
        rows.append({
            "Stock": clean_symbol(r.get("Stock")),
            "Signal": r.get("Signal"),
            "Score": r.get("Score"),
            "Close": r.get("Close"),
            "RSI": r.get("RSI"),
            "ADX": r.get("ADX"),
            "Supertrend": r.get("Supertrend"),
            "Weekly_Trend": r.get("Weekly_Trend"),
            "Entry_Low": r.get("Entry_Low"),
            "Entry_High": r.get("Entry_High"),
            "Stoploss": sl,
            "Target_1": t1,
            "Target_2": t2,
            "Final_Target": t3,
            "Support": r.get("Support"),
            "Resistance": r.get("Resistance"),
            "Risk_Reward": r.get("Risk_Reward"),
            "Patterns": ", ".join(r.get("Patterns", [])) if r.get("Patterns") else "",
        })
    return pd.DataFrame(rows)


def save_excel_report(ranked_result, path, glm_picks=None, breadth=None):
    """V8.3.0: Excel report with optional Market Breadth info row (top
    of Scan Result) + optional GLM Picks sheet."""
    # V8.2.0 FIX (bug #20): safe makedirs
    dirname = os.path.dirname(path) or "."
    os.makedirs(dirname, exist_ok=True)
    df = _prep_excel_dataframe(ranked_result)

    if df.empty:
        pd.DataFrame(columns=EXCEL_COLUMNS).to_excel(path, index=False)
        # Still add GLM Picks sheet if available (even when scan returned 0)
        if glm_picks:
            _add_glm_picks_sheet(path, glm_picks)
        return path

    df = df[EXCEL_COLUMNS]
    df.to_excel(path, index=False, sheet_name="Scan Result")

    wb = load_workbook(path)
    ws = wb["Scan Result"]

    # V8.3.0: Market Breadth info rows at top of "Scan Result" sheet
    if breadth:
        # Insert 2 blank rows at top (pushes headers to row 3, data to row 4+)
        ws.insert_rows(1, amount=2)
        # Row 1: sentiment summary (merged across all columns)
        sentiment = (breadth.get("sentiment") or "N/A").upper()
        sent_emoji = breadth.get("sentiment_emoji") or "🟡"
        ws.cell(row=1, column=1,
                value=f"MARKET SENTIMENT: {sentiment}  {sent_emoji}    |    "
                      f"Total Scanned: {breadth.get('total_scanned', 0)}    |    "
                      f"Advances: {breadth.get('advances', 0)}    |    "
                      f"Declines: {breadth.get('declines', 0)}    |    "
                      f"Breadth: {breadth.get('breadth_pct', 0)}%")
        ws.merge_cells(start_row=1, end_row=1,
                        start_column=1, end_column=len(EXCEL_COLUMNS))
        c1 = ws.cell(row=1, column=1)
        c1.fill = PatternFill("solid", fgColor="0D47A1")
        c1.font = Font(bold=True, color="FFFFFF", size=12)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        # Row 2: signal distribution (merged)
        ws.cell(row=2, column=1,
                value=f"Signal Distribution — 🔥 Strong Buy: {breadth.get('strong_buy', 0)}    "
                      f"| ⚡ Buy: {breadth.get('buy', 0)}    "
                      f"| 👀 Watch: {breadth.get('watch', 0)}    "
                      f"| ⚠️ Sell/Avoid: {breadth.get('sell_avoid', 0)}    "
                      f"| 🚀 New 20-day Highs: {breadth.get('new_highs', 0)}")
        ws.merge_cells(start_row=2, end_row=2,
                        start_column=1, end_column=len(EXCEL_COLUMNS))
        c2 = ws.cell(row=2, column=1)
        c2.fill = PatternFill("solid", fgColor="E3F2FD")
        c2.font = Font(bold=True, color="0D47A1", size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18
        header_row_idx = 3
        first_data_row = 4
        freeze_at = "A3"
    else:
        header_row_idx = 1
        first_data_row = 2
        freeze_at = "A2"

    # Header styling
    header_fill = PatternFill("solid", fgColor="263238")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Signal color-coding (per row)
    signal_col_idx = EXCEL_COLUMNS.index("Signal") + 1
    last_data_row = ws.max_row
    for row_idx in range(first_data_row, last_data_row + 1):
        signal_val = ws.cell(row=row_idx, column=signal_col_idx).value
        color = SIGNAL_COLORS.get(signal_val)
        if color:
            text_color = "FFFFFF" if signal_val in ("STRONG BUY", "SELL / AVOID") else "000000"
            cell = ws.cell(row=row_idx, column=signal_col_idx)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color=text_color)

    # Column widths
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        max_len = max(
            [len(str(col_name))]
            + [len(str(ws.cell(row=r, column=col_idx).value or ""))
               for r in range(first_data_row, last_data_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = freeze_at

    # Top Buy List sheet (existing behavior)
    top_df = df[df["Signal"].isin(["STRONG BUY", "BUY"])]
    if not top_df.empty:
        ws2 = wb.create_sheet("Top Buy List")
        ws2.append(["Rank", "Stock", "Signal", "Score", "Entry_Low", "Entry_High",
                    "Stoploss", "Target_1", "Target_2", "Final_Target"])
        for c in range(1, 11):
            cell = ws2.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
        for i, (_, r) in enumerate(top_df.iterrows(), start=1):
            ws2.append([i, r["Stock"], r["Signal"], r["Score"], r["Entry_Low"],
                        r["Entry_High"], r["Stoploss"], r["Target_1"], r["Target_2"],
                        r["Final_Target"]])
        ws2.freeze_panes = "A2"

    # V8.3.0: GLM Picks sheet (if available)
    if glm_picks:
        _add_glm_picks_sheet_to_wb(wb, glm_picks)

    wb.save(path)
    return path


def _add_glm_picks_sheet_to_wb(wb, glm_picks):
    """V8.3.0: 'GLM Picks' sheet banata hai — rank, stock, name,
    confidence, action, rationale, risk note, tech score, tech signal."""
    ws3 = wb.create_sheet("GLM Picks")
    glm_headers = ["Rank", "Stock", "Name", "Confidence", "Action",
                   "Rationale", "Risk Note", "Tech Score", "Tech Signal"]
    ws3.append(glm_headers)
    header_fill = PatternFill("solid", fgColor="0D47A1")
    header_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(glm_headers) + 1):
        cell = ws3.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for p in glm_picks:
        ws3.append([
            p.get("rank"),
            clean_symbol(p.get("stock", "")),
            p.get("name", ""),
            p.get("glm_confidence"),
            p.get("glm_action"),
            p.get("glm_rationale", ""),
            p.get("glm_risk_note", ""),
            p.get("technical_score"),
            p.get("technical_signal", ""),
        ])
    # Column widths
    col_widths = [6, 14, 24, 12, 14, 50, 40, 11, 14]
    for col_idx, w in enumerate(col_widths, start=1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    # Color Action column
    action_col_idx = glm_headers.index("Action") + 1
    for row_idx in range(2, ws3.max_row + 1):
        action_val = ws3.cell(row=row_idx, column=action_col_idx).value
        color = GLM_ACTION_COLORS.get(action_val)
        if color:
            cell = ws3.cell(row=row_idx, column=action_col_idx)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True)
        # Wrap rationale/risk note text
        ws3.cell(row=row_idx, column=glm_headers.index("Rationale") + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(row=row_idx, column=glm_headers.index("Risk Note") + 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.freeze_panes = "A2"


def _add_glm_picks_sheet(path, glm_picks):
    """Helper: open existing xlsx and add GLM Picks sheet (used when
    ranked_result was empty but glm_picks is available — edge case)."""
    try:
        wb = load_workbook(path)
        _add_glm_picks_sheet_to_wb(wb, glm_picks)
        wb.save(path)
    except Exception as e:
        logger.warning(f"GLM Picks sheet add fail: {e}")


# ---------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------
def save_report(ranked_result, chart_output_data, excel_path="reports/AI_Report.xlsx", pdf_path="reports/AI_Report.pdf", glm_picks=None, breadth=None):
    """
    V8.3.0: glm_picks + breadth optional params (G2 wiring-integrator
    ne add kiye). Ye values ab Excel (breadth info row + GLM Picks
    sheet) + PDF (Market Sentiment section + GLM AI Top Picks section
    + per-stock GLM View subsection) dono mein use hote hain.

    Existing call signatures backward-compatible (defaults None).
    """
    if isinstance(chart_output_data, tuple):
        chart_paths = chart_output_data[0]
    else:
        chart_paths = chart_output_data or {}

    # V8.2.0 FIX (bug #20): safe makedirs (handles bare filename paths).
    excel_dir = os.path.dirname(excel_path) or "."
    os.makedirs(excel_dir, exist_ok=True)
    # V8.3.0: pass glm_picks + breadth to Excel for GLM Picks sheet + breadth info row
    save_excel_report(ranked_result, excel_path, glm_picks=glm_picks, breadth=breadth)

    top_stocks_data = []
    top_buys = [r for r in ranked_result if r.get("Signal") in ("STRONG BUY", "BUY")]

    for r in top_buys[:10]:
        stock = r.get("Stock")
        entry = r.get("Entry")
        sl = r.get("Stoploss")
        t1, t2, t3 = calculate_targets(entry, sl) if (entry and sl) else (0.0, 0.0, 0.0)

        top_stocks_data.append({
            'symbol': clean_symbol(stock),
            'score': r.get("Score", 0),
            'signal': r.get("Signal", ""),
            'entry': entry or 0.0,
            'entry_low': r.get("Entry_Low"),
            'entry_high': r.get("Entry_High"),
            'sl': sl or 0.0,
            't1': t1 or 0.0,
            't2': t2 or 0.0,
            't3': t3 or 0.0,
            'support': r.get("Support") or 0.0,
            'resistance': r.get("Resistance") or 0.0,
            'pattern': ", ".join(r.get("Patterns", [])) if r.get("Patterns") else "None",
            'weekly_trend': r.get("Weekly_Trend", "N/A"),
            'mtf_status': r.get("MTF_1H_Status", "N/A"),
            # V8.3.0: extra fields for richer card header
            'close': r.get("Close", 0.0) or 0.0,
            'rsi': r.get("RSI", "—"),
            'adx': r.get("ADX", "—"),
            'risk_reward': r.get("Risk_Reward", "—"),
            # REAL AI analysis (scanner.py se) - boilerplate nahi
            'ai_analysis': r.get("AI_Analysis", "Analysis available nahi hai."),
            'news': format_news_text(stock, limit=2),
            'chart_path': chart_paths.get(stock, ""),
        })

    # V8.3.0: PDF mein glm_picks + breadth use hote hain (Market
    # Sentiment + GLM AI Top Picks + per-stock GLM View subsection).
    generate_pdf_report(top_stocks_data, pdf_path, glm_picks=glm_picks, breadth=breadth)
    return excel_path, pdf_path
